#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Módulo de leitura de Excel para o aplicativo Wall Engenharia.
Responsável por extrair dados das planilhas de orçamento.
"""

import os
import logging
from typing import Dict, List, Any, Optional, Union
import pandas as pd
from openpyxl import load_workbook

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('excel_reader')

class ExcelReader:
    """Classe para leitura e extração de dados de planilhas Excel."""
    
    def __init__(self, file_path: str):
        """
        Inicializa o leitor de Excel.
        
        Args:
            file_path: Caminho completo para o arquivo Excel.
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet_names = []
        self.current_sheet = None
        self.data = {}
        
        # Verificar se o arquivo existe
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Arquivo Excel não encontrado: {file_path}")
        
        # Carregar o arquivo
        try:
            self.workbook = load_workbook(file_path, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            logger.info(f"Arquivo Excel carregado com sucesso: {file_path}")
            logger.info(f"Planilhas disponíveis: {self.sheet_names}")
        except Exception as e:
            logger.error(f"Erro ao carregar arquivo Excel: {e}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """Retorna a lista de nomes das planilhas disponíveis."""
        return self.sheet_names
    
    def select_sheet(self, sheet_name: str) -> bool:
        """
        Seleciona uma planilha para trabalhar.
        
        Args:
            sheet_name: Nome da planilha a ser selecionada.
            
        Returns:
            bool: True se a planilha foi selecionada com sucesso, False caso contrário.
        """
        if sheet_name in self.sheet_names:
            self.current_sheet = self.workbook[sheet_name]
            logger.info(f"Planilha selecionada: {sheet_name}")
            return True
        else:
            logger.warning(f"Planilha não encontrada: {sheet_name}")
            return False
    
    def extract_cell_value(self, row: int, col: int) -> Any:
        """
        Extrai o valor de uma célula específica.
        
        Args:
            row: Número da linha (1-based).
            col: Número da coluna (1-based).
            
        Returns:
            Valor da célula.
        """
        if not self.current_sheet:
            logger.error("Nenhuma planilha selecionada")
            return None
        
        try:
            return self.current_sheet.cell(row=row, column=col).value
        except Exception as e:
            logger.error(f"Erro ao extrair valor da célula ({row}, {col}): {e}")
            return None
    
    def find_cell_by_value(self, search_value: str, partial_match: bool = False) -> Optional[tuple]:
        """
        Procura uma célula pelo seu valor.
        
        Args:
            search_value: Valor a ser procurado.
            partial_match: Se True, procura por correspondência parcial.
            
        Returns:
            Tupla (linha, coluna) da célula encontrada ou None se não encontrada.
        """
        if not self.current_sheet:
            logger.error("Nenhuma planilha selecionada")
            return None
        
        for row in range(1, self.current_sheet.max_row + 1):
            for col in range(1, self.current_sheet.max_column + 1):
                cell_value = self.current_sheet.cell(row=row, column=col).value
                
                if cell_value:
                    if isinstance(cell_value, str):
                        if partial_match and search_value.lower() in cell_value.lower():
                            return (row, col)
                        elif not partial_match and search_value.lower() == cell_value.lower():
                            return (row, col)
        
        logger.warning(f"Valor '{search_value}' não encontrado na planilha")
        return None
    
    def extract_data_for_proposal(self) -> Dict[str, Any]:
        """
        Extrai os dados necessários para a proposta.
        
        Returns:
            Dicionário com os dados extraídos.
        """
        if not self.current_sheet:
            logger.error("Nenhuma planilha selecionada")
            return {}
        
        data = {
            'nome_cliente': '',
            'nome_contato': '',
            'email': '',
            'telefone': '',
            'escopo': '',
            'prazo': '',
            'custo': None,
            'garantias': '',
            'seguro': '',
            'nao_inclusos': ''
        }
        
        # Extrair nome do cliente (geralmente nas primeiras linhas)
        for row in range(1, 10):
            for col in range(1, 10):
                cell_value = self.extract_cell_value(row, col)
                if cell_value and isinstance(cell_value, str) and len(cell_value) > 3:
                    data['nome_cliente'] = cell_value
                    break
        
        # Procurar por valores de custo (geralmente nas linhas finais)
        for row in range(self.current_sheet.max_row - 20, self.current_sheet.max_row + 1):
            for col in range(1, self.current_sheet.max_column + 1):
                cell_label = self.extract_cell_value(row, 2)  # Coluna B geralmente contém os rótulos
                cell_value = self.extract_cell_value(row, col)
                
                if cell_label and isinstance(cell_label, str) and "CUSTO FINAL" in cell_label:
                    if cell_value and isinstance(cell_value, (int, float)) and cell_value > 1000:
                        data['custo'] = cell_value
                        break
        
        # Procurar por informações de seguro
        seguro_pos = self.find_cell_by_value("SEGURO", partial_match=True)
        if seguro_pos:
            row, col = seguro_pos
            data['seguro'] = self.extract_cell_value(row, col)
        
        logger.info("Dados extraídos da planilha Excel")
        return data
    
    def extract_all_scenarios(self) -> Dict[str, Dict[str, Any]]:
        """
        Extrai dados de todos os cenários (planilhas) disponíveis.
        
        Returns:
            Dicionário com os dados de cada cenário.
        """
        scenarios = {}
        
        for sheet_name in self.sheet_names:
            self.select_sheet(sheet_name)
            scenarios[sheet_name] = self.extract_data_for_proposal()
        
        return scenarios
    
    def close(self):
        """Fecha o arquivo Excel."""
        if self.workbook:
            self.workbook.close()
            logger.info("Arquivo Excel fechado")


# Função auxiliar para uso direto
def extract_data_from_excel(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Extrai dados de um arquivo Excel para uso na proposta.
    
    Args:
        file_path: Caminho para o arquivo Excel.
        sheet_name: Nome da planilha específica (opcional).
        
    Returns:
        Dicionário com os dados extraídos.
    """
    reader = ExcelReader(file_path)
    
    if sheet_name:
        if reader.select_sheet(sheet_name):
            data = reader.extract_data_for_proposal()
        else:
            data = {}
    else:
        # Se nenhuma planilha for especificada, usa a primeira
        if reader.sheet_names:
            reader.select_sheet(reader.sheet_names[0])
            data = reader.extract_data_for_proposal()
        else:
            data = {}
    
    reader.close()
    return data


if __name__ == "__main__":
    # Teste básico do módulo
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = input("Digite o caminho para o arquivo Excel: ")
    
    try:
        reader = ExcelReader(file_path)
        print(f"Planilhas disponíveis: {reader.get_sheet_names()}")
        
        for sheet_name in reader.get_sheet_names():
            print(f"\nExtraindo dados da planilha: {sheet_name}")
            reader.select_sheet(sheet_name)
            data = reader.extract_data_for_proposal()
            
            print("Dados extraídos:")
            for key, value in data.items():
                if value:  # Mostrar apenas campos com valor
                    print(f"  {key}: {value}")
        
        reader.close()
    except Exception as e:
        print(f"Erro: {e}")
        sys.exit(1)
